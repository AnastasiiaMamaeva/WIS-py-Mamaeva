import argparse
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Define NCBI API base URLs
SEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
FETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"

def search_ncbi(database, term, max_results):
    """Search the NCBI database and return a list of IDs and total count."""
    params = {
        "db": database,
        "term": term,
        "retmax": max_results,
        "retmode": "json"
    }
    response = requests.get(SEARCH_URL, params=params)
    response.raise_for_status()
    data = response.json()
    ids = data.get("esearchresult", {}).get("idlist", [])
    total_count = int(data.get("esearchresult", {}).get("count", 0))
    return ids, total_count

def fetch_ncbi_item(database, item_id):
    """Fetch an item from the NCBI database by its ID."""
    params = {
        "db": database,
        "id": item_id,
        "rettype": "gb",
        "retmode": "text"
    }
    response = requests.get(FETCH_URL, params=params)
    response.raise_for_status()
    return response.text

def save_log_to_excel(log_file, date, database, term, max_results, total_found):
    """Save log data to an Excel file."""
    try:
        # Load workbook if it exists, otherwise create a new one
        try:
            workbook = load_workbook(log_file)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            # Write headers if the workbook is new
            sheet.append(["date", "database", "term", "max", "total"])

        # Append the log entry
        sheet.append([date, database, term, max_results, total_found])
        workbook.save(log_file)
    except Exception as e:
        print(f"Error saving log to Excel: {e}")

def main():
    # Define command-line arguments
    parser = argparse.ArgumentParser(description="Download data from NCBI.")
    parser.add_argument("--database", default="nucleotide", help="NCBI database to search (default: nucleotide)")
    parser.add_argument("--term", required=True, help="Search term")
    parser.add_argument("--number", default=10, type=int, help="Number of items to download (default: 10)")
    parser.add_argument("--logfile", default="ncbi_log.xlsx", help="Log file to store metadata (default: ncbi_log.xlsx)")

    args = parser.parse_args()
    database = args.database
    term = args.term
    max_results = args.number
    log_file = args.logfile

    # Perform search
    print(f"Searching {database} for '{term}'...")
    try:
        ids, total_found = search_ncbi(database, term, max_results)
    except Exception as e:
        print(f"Error during search: {e}")
        return

    # Download items
    print(f"Found {total_found} items. Downloading up to {max_results} items...")
    downloaded_files = []
    for idx, item_id in enumerate(ids):
        try:
            data = fetch_ncbi_item(database, item_id)
            filename = f"{term.replace(' ', '_')}_{idx + 1}.txt"
            with open(filename, "w") as file:
                file.write(data)
            downloaded_files.append(filename)
        except Exception as e:
            print(f"Error downloading item {item_id}: {e}")
            continue

    # Print downloaded files
    print("Downloaded files:")
    for file in downloaded_files:
        print(f"- {file}")

    # Save metadata to the Excel log
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_log_to_excel(log_file, current_date, database, term, max_results, total_found)
    print(f"Metadata saved in '{log_file}'.")

if __name__ == "__main__":
    main()
